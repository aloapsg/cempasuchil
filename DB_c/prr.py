#PRUEBA_1 EXTRAER CONVERSACIÓN DE CHAT.TXT Y DATOS P DE XLS
from pymongo import MongoClient
import os
import nltk
import openpyxl


mongoClient = MongoClient('localhost',27017)
db = mongoClient.PRUEBA_2
collection = db.conver

# DATOS PERSONALES

doc = openpyxl.load_workbook('Plantilla_while.xlsx')
names_sheets=doc.get_sheet_names()
sheet = doc.get_sheet_by_name('Hoja 1')

mujer='M'
hombre='H'

#diccionario[clave] = valor

conversation={} #diccionario que contiene todo DAD
usuarios=[] # no recuerdo
rw=2
cn=1
conversaciones = []

def convers(j,k):
    users=[]
    dic={}
    usuario1={} #luego usuario1 va a dic y dic a conversation
    usuario2={}
    id=""
    conv1=""
    conv2=""
    if str(sheet.cell(row=j,column=1).value).find(".txt") != -1:
        id = str(sheet.cell(row=j,column=1).value).replace(".txt","")
    else:
        id=sheet.cell(row=j,column=1).value

    arch = open(id + ".txt","r");
    lines = arch.readlines()
    texto=""

    users.append(str(sheet.cell(row=j,column=2).value))
    users.append(str(sheet.cell(row=k,column=2).value))
    print(users)

    for line in lines:
        text=line
        texto = texto + text
    arch.close()

    for user in users:
        if users.index(user) == 0:
            texto=texto.replace(user,"USER1")
        else:
            texto=texto.replace(user,"USER2")

    lineas = texto.splitlines(3)
    for linea in lineas:
        if linea.find("USER1") != -1:
            conv1 = conv1 + linea
        elif linea.find("USER2") != -1:
            conv2 = conv2 + linea
    
    def user1(j,a):
        if str(sheet.cell(row=j,column=1).value).find(".txt") != -1:
            id1= str(sheet.cell(row=j,column=1).value).replace(".txt","")
        else:
            id1=sheet.cell(row=j,column=1).value

        print(id1)
        dic["ID"]=id1
        
        usuario1['age']=sheet.cell(row=j,column=4).value
        usuario1['gender'] = sheet.cell(row=j,column=5).value
        usuario1['osexual']= sheet.cell(row=j,column=6).value  #genero / gender

        usuario1['birthPlace'] = sheet.cell(row=j,column=7).value
        usuario1['livePlace'] = sheet.cell(row=j,column=8).value
        usuario1['languages'] = sheet.cell(row=j,column=9).value
        usuario1['education'] = sheet.cell(row=j,column=10).value
        usuario1['faculty'] = sheet.cell(row=j,column=11).value
        usuario1['major'] = sheet.cell(row=j,column=12).value
        usuario1['ocupation'] = sheet.cell(row=j,column=13).value
        usuario1['prof_of'] = sheet.cell(row=j,column=14).value
        usuario1['relation'] = sheet.cell(row=j,column=15).value
        usuario1['contact'] = sheet.cell(row=j,column=16).value
        usuario1['email'] = sheet.cell(row=j,column=17).value
        usuario1['lines'] = a
        return usuario1

    def user2(k,b):
        usuario2['age']=sheet.cell(row=k,column=4).value
        usuario2['gender'] = sheet.cell(row=k,column=5).value
        usuario2['osexual']= sheet.cell(row=k,column=6).value  #genero / gender
        
        usuario2['birthPlace'] = sheet.cell(row=k,column=7).value
        usuario2['livePlace'] = sheet.cell(row=k,column=8).value
        usuario2['languages'] = sheet.cell(row=k,column=9).value
        usuario2['education'] = sheet.cell(row=k,column=10).value
        usuario2['faculty'] = sheet.cell(row=k,column=11).value
        usuario2['major'] = sheet.cell(row=k,column=12).value
        usuario2['ocupation'] = sheet.cell(row=k,column=13).value
        usuario2['prof_of'] = sheet.cell(row=k,column=14).value
        usuario2['relation'] = sheet.cell(row=k,column=15).value
        usuario2['lines'] = b
        return usuario2

    user1(j,conv1)
    user2(k,conv2)
    dic["user1"]=usuario1
    dic["user2"]=usuario2
    dic["conversacion"]=texto
    return dic

#acaba func que extrae datos 146
cont=0
for reng in range(2,150):
    reng+=cont
    conversaciones.append(convers(reng,reng+1))
    cont+=1

#print(conversaciones)
for elm in conversaciones:
 collection.insert(elm)
