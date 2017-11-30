#PRUEBA_1 EXTRAER CONVERSACIÓN DE CHAT.TXT Y DATOS P DE XLS
from pymongo import MongoClient
import os
import nltk
import openpyxl
import pprint


mongoClient = MongoClient('localhost',27017)
db = mongoClient.PRUEBA_2
collection = db.conver

# Obtener datos
doc = openpyxl.load_workbook('Plantilla_5.xlsx')
names_sheets=doc.get_sheet_names()
sheet = doc.get_sheet_by_name('Hoja 1')

todos_conver = []
directorio = {}

inicio = 0 #818
cont = 1
for reng in range(2,818):
    res1 = ""
    res2 = ""
    reng+=inicio
    num1 = sheet.cell(row=reng,column=3).value
    num2 = sheet.cell(row=reng+1,column=3).value
    
    for num in directorio:
        if(num1 == num):
            res1 = "ya esta"
        if(num2 == num):
            res2 = "ya esta"

    if(res1 != "ya esta"):
        directorio[num1]= "USER" + str(cont)
        cont+=1
    if(res2 != "ya esta"):
        directorio[num2]= "USER" + str(cont)
        cont+=1
    inicio+=1

Familiar = ["Hija", "Hijo", "Madre", "Padre", "Hermano", "Hermana", "Cuñada", "Cuñado", "Prima", "Primo", "Sobrino", "Sobrina", "Tío", "Tía"]
Amistad = ["Amigo","Amiga"]
Amorosa = ["Novia","Novio"]
Laboral = ["Empleada", "Empleado", "Jefa", "Jefe", "Socio", "Socia", "Compañera De Trabajo", "Compañero De Trabajo"]
Academica = ["Alumno", "Alumna", "Profesor", "Profesora", "Compañera De Escuela", "Compañero De Escuela"]
Conocidos = ["Conocida","Conocido","Ex-novia","Ex-novio"]
Profesional = ["Paciente", "Terapeuta"]


def paquete(reng1,reng2):
    
    conversacion = {}
    us1 = []
    us2 = []
    user1 = {}
    user2 = {}
    relacion = ""
    texto = ""
    
    id = sheet.cell(row=reng1,column=1).value
    arch = open(id ,"r");
    lines= arch.readlines()
    
    num1 = sheet.cell(row=reng1,column=3).value
    num2 = sheet.cell(row=reng2,column=3).value
    
    us1.append(directorio[num1])
    us2.append(directorio[num2])
    
    for col in range(4,16):
        us1.append(sheet.cell(row=reng1,column=col).value)
        us2.append(sheet.cell(row=reng2,column=col).value)
    
    user1['user'] = us1[0]
    user1['age'] = us1[1]
    user1['gender'] = us1[2]
    user1['oSexual'] = us1[3]
    user1['bPlace'] = us1[4]
    user1['lPlace'] = us1[5]
    user1['lang'] = us1[6]
    user1['edu'] = us1[7]
    user1['fac'] = us1[8]
    user1['major'] = us1[9]
    user1['ocup'] = us1[10]
    user1['prof_of'] = us1[11]
    user1['rel'] = us1[12]
        
    user2['user'] = us2[0]
    user2['age'] = us2[1]
    user2['gender'] = us2[2]
    user2['oSexual'] = us2[3]
    user2['bPlace'] = us2[4]
    user2['lPlace'] = us2[5]
    user2['lang'] = us2[6]
    user2['edu'] = us2[7]
    user2['fac'] = us2[8]
    user2['major'] = us2[9]
    user2['ocup'] = us2[10]
    user2['prof_of'] = us2[11]
    user2['rel'] = us2[12]

    for a in Familiar:
        if(user1['rel'] == a):
            relacion = "Familiar"

    for a in Amistad:
        if(user1['rel'] == a):
            relacion = "Amistad"

    for a in Amorosa:
        if(user1['rel'] == a):
            relacion = "Amorosa"

    for a in Laboral:
        if(user1['rel'] == a):
            relacion = "Laboral"

    for a in Academica:
        if(user1['rel'] == a):
            relacion = "Academica"

    for a in Conocidos:
        if(user1['rel'] == a):
            relacion = "Conocidos"

    for a in Profesional:
        if(user1['rel'] == a):
            relacion = "Profesional"


    for line in lines:
        text=line
        texto = texto + text

    conversacion['user1'] = user1
    conversacion['user2'] = user2
    conversacion['conv'] = texto
    conversacion['relation'] = relacion
              
    return conversacion

inic = 0
for reng in range(2,818):
    reng+=inic
    todos_conver.append(paquete(reng,reng+1))
    inic+=1

for elm in todos_conver:
    collection.insert(elm)
